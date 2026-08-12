"""Where-is-the-table detection, kept separate from the mechanical cell-reading
in ``xlsx_reader``/``csv_reader``.

This is the extension point for the hard, open-ended part of principle 5
(offset headers, merged cells, multiple tables per sheet). ``detect_header_row``
tackles the offset-header case — title rows and blank rows above the real
header, common in exported reports — by scoring nearby rows on how header-like
they look relative to what's beneath them, instead of assuming row 0. Merged
cells and multiple tables per sheet remain open; ``AutoDetectHeader`` still
treats the sheet's used range as exactly one table.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Protocol, Sequence

from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class DetectedTable:
    """A rectangular region of a worksheet identified as one table, in
    0-based, end-exclusive coordinates."""

    header_row: int
    start_col: int
    end_col: int
    data_start_row: int
    data_end_row: int


class TableDetectionStrategy(Protocol):
    def detect(self, sheet: Worksheet) -> list[DetectedTable]: ...


def _is_blank_row(row: Sequence[Any]) -> bool:
    return all(v is None or v == "" for v in row)


def _type_category(value: Any) -> str:
    """Coarse type bucket used to compare a candidate header cell against the
    data beneath it. Numeric-looking strings (as read from CSV) count as
    numbers so the same heuristic works for both readers."""
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (date, datetime)):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        try:
            float(value.strip())
            return "number"
        except ValueError:
            return "text"
    return "text"


def _header_score(candidate: Sequence[Any], sample_rows: list[Sequence[Any]]) -> float:
    """How much `candidate` looks like a header for the rows beneath it.

    A real header is mostly filled in, and each column's label tends to be
    text even when the data in that column below is typed (numbers, dates,
    booleans) — that contrast is the main signal. A fully blank row can never
    be a header. Missing individual cells (e.g. one unnamed column) are
    tolerated rather than disqualifying, since that's a normal, already
    supported shape (columns fall back to "Coluna N")."""
    filled = [v for v in candidate if v is not None and v != ""]
    if not filled:
        return -1.0
    if not sample_rows:
        return 0.0

    matches = 0
    considered = 0
    for col_idx, cand_value in enumerate(candidate):
        if cand_value is None or cand_value == "":
            continue
        col_values = [
            row[col_idx] for row in sample_rows if col_idx < len(row) and row[col_idx] not in (None, "")
        ]
        if not col_values:
            continue
        considered += 1
        counts: dict[str, int] = {}
        for v in col_values:
            category = _type_category(v)
            counts[category] = counts.get(category, 0) + 1
        majority_category = max(counts, key=counts.get)
        if _type_category(cand_value) != majority_category:
            matches += 1

    match_ratio = matches / considered if considered else 0.0
    fill_ratio = len(filled) / len(candidate)
    return match_ratio * fill_ratio


def detect_header_row(rows: Sequence[Sequence[Any]], *, window: int = 10, sample_size: int = 10) -> int:
    """Best-effort guess at which row in `rows` is the header.

    Skips fully blank leading rows, then scores nearby candidates (title
    rows, the header itself, and the first couple of data rows) by how
    header-like they look relative to a sample of the rows beneath them, and
    returns the top scorer. Ties favor the earliest row. Falls back to the
    first non-blank row when nothing scores convincingly, so behavior
    degrades to the old top-of-sheet assumption rather than guessing wildly
    on unfamiliar shapes.

    Returns 0 for an empty `rows`.
    """
    first_non_blank = next((i for i, row in enumerate(rows) if not _is_blank_row(row)), 0)
    if first_non_blank >= len(rows):
        return first_non_blank

    best_idx = first_non_blank
    best_score = float("-inf")
    for idx in range(first_non_blank, min(first_non_blank + window, len(rows))):
        sample = rows[idx + 1 : idx + 1 + sample_size]
        score = _header_score(rows[idx], sample)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


class SingleTableFromTopLeft:
    """Literal baseline strategy: the sheet's used range is one table, header
    always on its first row — no attempt to skip title/blank rows. Kept for
    callers that want that exact assumption; ``AutoDetectHeader`` is the
    smarter default. Does not attempt to handle merged cells or multiple
    tables — see module docstring."""

    def detect(self, sheet: Worksheet) -> list[DetectedTable]:
        if sheet.max_row < 1 or sheet.max_column < 1:
            return []
        return [
            DetectedTable(
                header_row=0,
                start_col=0,
                end_col=sheet.max_column,
                data_start_row=1,
                data_end_row=sheet.max_row,
            )
        ]


class AutoDetectHeader:
    """Default strategy: the sheet's used range is one table, but the header
    row is picked with ``detect_header_row`` instead of assumed to be row 0 —
    handles leading blank rows and a title row above the real header. Still
    doesn't split multiple tables per sheet or handle merged cells — see
    module docstring."""

    def detect(self, sheet: Worksheet) -> list[DetectedTable]:
        if sheet.max_row < 1 or sheet.max_column < 1:
            return []
        grid = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True))
        header_row = detect_header_row(grid)
        return [
            DetectedTable(
                header_row=header_row,
                start_col=0,
                end_col=sheet.max_column,
                data_start_row=header_row + 1,
                data_end_row=sheet.max_row,
            )
        ]
