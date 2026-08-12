from __future__ import annotations

from pathlib import Path

import openpyxl

from tabula_engine.canonical.model import CanonicalColumn, CanonicalTable, SourceRef
from .detection import AutoDetectHeader, TableDetectionStrategy
from .type_inference import infer_column_type


class XlsxReader:
    """Cell-by-cell .xlsx reader. Detection (where the table is) and reading
    (what's in it) are split so a smarter ``TableDetectionStrategy`` can be
    swapped in later without touching this class."""

    def __init__(self, detection_strategy: TableDetectionStrategy | None = None):
        self._detection_strategy = detection_strategy or AutoDetectHeader()

    def read(self, path: str | Path) -> list[CanonicalTable]:
        path = Path(path)
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        tables: list[CanonicalTable] = []
        try:
            for sheet in workbook.worksheets:
                detected = self._detection_strategy.detect(sheet)
                for table_index, region in enumerate(detected):
                    tables.append(self._read_region(path.name, sheet, table_index, region))
        finally:
            workbook.close()
        return tables

    def _read_region(self, file_name: str, sheet, table_index: int, region) -> CanonicalTable:
        header_cells = list(
            sheet.iter_rows(
                min_row=region.header_row + 1,
                max_row=region.header_row + 1,
                min_col=region.start_col + 1,
                max_col=region.end_col,
                values_only=True,
            )
        )[0]
        column_names = [str(v) if v not in (None, "") else f"Coluna {i + 1}" for i, v in enumerate(header_cells)]

        raw_rows = list(
            sheet.iter_rows(
                min_row=region.data_start_row + 1,
                max_row=region.data_end_row,
                min_col=region.start_col + 1,
                max_col=region.end_col,
                values_only=True,
            )
        )

        columns: list[CanonicalColumn] = []
        for col_idx, name in enumerate(column_names):
            sample = [r[col_idx] for r in raw_rows]
            columns.append(
                CanonicalColumn(
                    name=name,
                    type=infer_column_type(sample),
                    source_column=region.start_col + col_idx,
                )
            )

        rows = [{col.name: raw_row[i] for i, col in enumerate(columns)} for raw_row in raw_rows]

        source = SourceRef(
            file_name=file_name,
            sheet_name=sheet.title,
            table_index=table_index,
            header_row=region.header_row,
        )
        return CanonicalTable(
            columns=columns,
            rows=rows,
            source=source,
            row_source_offset=region.data_start_row,
        )
